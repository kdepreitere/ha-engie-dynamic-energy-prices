"""Data coordinator for Engie Dynamic Prices."""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Optional

from homeassistant.core import HomeAssistant
from homeassistant.helpers.update_coordinator import DataUpdateCoordinator, UpdateFailed

from .const import (
    CONF_BUY_DISTRIBUTION_FEE,
    CONF_BUY_ENERGY_FEE,
    CONF_BUY_MULTIPLIER,
    CONF_SELL_FEE,
    DOMAIN,
    SELL_ALERT_THRESHOLD,
    XLSX_URL,
)

_LOGGER = logging.getLogger(__name__)

# Coordinator polls every 15 min so the 15-min price sensor stays current.
# Actual HTTP fetches only happen when new data is needed.
SCAN_INTERVAL = timedelta(minutes=15)

# Day-ahead prices published by EPEX at ~14:00.
AFTERNOON_HOUR = 14

# Number of 15-min slots per day.
SLOTS_15MIN = 96


def _to_kwh(eur_mwh: float) -> float:
    """Convert EUR/MWh to EUR/kWh."""
    return round(float(eur_mwh) / 1000, 6)


def _hour_label(i: int) -> str:
    return f"{i:02d}:00 - {(i + 1) % 24:02d}:00"


def _quarter_label(i: int) -> str:
    start = i * 15
    end = (i + 1) * 15
    sh, sm = divmod(start, 60)
    eh, em = divmod(end % (24 * 60), 60)
    return f"{sh:02d}:{sm:02d} - {eh:02d}:{em:02d}"


def _to_hourly(prices_15min: list[float]) -> list[float]:
    """Average every 4 consecutive quarter-hour slots into one hourly value."""
    return [
        round(sum(prices_15min[i * 4:(i + 1) * 4]) / 4, 6)
        for i in range(24)
        if len(prices_15min) >= (i + 1) * 4
    ]


@dataclass
class EngieData:
    """Parsed daily price data from Engie."""

    price_date: date

    # Spot prices (EUR/kWh) -- raw EPEX values before any formula is applied.
    prices: list[float]                # today:     24 hourly values (derived from 15-min)
    prices_tomorrow: list[float]       # tomorrow:  24 hourly values (after ~14:00)
    prices_15min: list[float]          # today:     96 quarter-hour values
    prices_15min_tomorrow: list[float] # tomorrow:  96 quarter-hour values (after ~14:00)

    # Contract formula parameters (configured by the user).
    #   buy  = (buy_multiplier * spot + buy_energy_fee) + buy_distribution_fee
    #   sell = spot - sell_fee
    buy_multiplier: float = 1.0
    buy_energy_fee: float = 0.0
    buy_distribution_fee: float = 0.0
    sell_fee: float = 0.0

    # ------------------------------------------------------------------ helpers
    def _buy(self, spot: float) -> float:
        return round(
            (self.buy_multiplier * spot + self.buy_energy_fee) + self.buy_distribution_fee,
            6,
        )

    def _sell(self, spot: float) -> float:
        return round(spot - self.sell_fee, 6)

    # ------------------------------------------------------- today -- spot hourly
    @property
    def current_price(self) -> float | None:
        h = datetime.now().hour
        return self.prices[h] if len(self.prices) > h else None

    @property
    def next_price(self) -> float | None:
        h = (datetime.now().hour + 1) % 24
        return self.prices[h] if len(self.prices) > h else None

    @property
    def average_price(self) -> float | None:
        return round(sum(self.prices) / len(self.prices), 6) if self.prices else None

    @property
    def min_price(self) -> float | None:
        src = self.prices_15min or self.prices
        return min(src) if src else None

    @property
    def max_price(self) -> float | None:
        src = self.prices_15min or self.prices
        return max(src) if src else None

    @property
    def cheapest_hours(self) -> list[dict]:
        return [
            {"hour": _hour_label(i), "price_eur_kwh": p}
            for i, p in sorted(enumerate(self.prices), key=lambda x: x[1])
        ]

    @property
    def all_prices_today(self) -> list[dict]:
        return [{"hour": _hour_label(i), "price_eur_kwh": p} for i, p in enumerate(self.prices)]

    @property
    def negative_price_hours_today(self) -> list[str]:
        return [_hour_label(i) for i, p in enumerate(self.prices) if p < 0]

    # ------------------------------------------------------- today -- buy (contract)
    @property
    def current_contract_price(self) -> float | None:
        p = self.current_price
        return self._buy(p) if p is not None else None

    @property
    def next_contract_price(self) -> float | None:
        p = self.next_price
        return self._buy(p) if p is not None else None

    @property
    def average_contract_price(self) -> float | None:
        if not self.prices:
            return None
        return round(sum(self._buy(p) for p in self.prices) / len(self.prices), 6)

    @property
    def min_contract_price(self) -> float | None:
        src = self.prices_15min or self.prices
        return self._buy(min(src)) if src else None

    @property
    def max_contract_price(self) -> float | None:
        src = self.prices_15min or self.prices
        return self._buy(max(src)) if src else None

    # ------------------------------------------------------- today -- sell
    @property
    def current_sell_price(self) -> float | None:
        p = self.current_price
        return self._sell(p) if p is not None else None

    @property
    def next_sell_price(self) -> float | None:
        p = self.next_price
        return self._sell(p) if p is not None else None

    @property
    def average_sell_price(self) -> float | None:
        if not self.prices:
            return None
        return round(sum(self._sell(p) for p in self.prices) / len(self.prices), 6)

    @property
    def min_sell_price(self) -> float | None:
        src = self.prices_15min or self.prices
        return self._sell(min(src)) if src else None

    @property
    def max_sell_price(self) -> float | None:
        src = self.prices_15min or self.prices
        return self._sell(max(src)) if src else None

    @property
    def negative_sell_hours_today(self) -> list[str]:
        return [_hour_label(i) for i, p in enumerate(self.prices) if self._sell(p) < SELL_ALERT_THRESHOLD]

    # ------------------------------------------------------- today -- 15-min spot
    @property
    def current_price_15min(self) -> float | None:
        now = datetime.now()
        slot = now.hour * 4 + now.minute // 15
        return self.prices_15min[slot] if len(self.prices_15min) > slot else None

    @property
    def next_price_15min(self) -> float | None:
        now = datetime.now()
        slot = (now.hour * 4 + now.minute // 15 + 1) % SLOTS_15MIN
        return self.prices_15min[slot] if len(self.prices_15min) > slot else None

    @property
    def average_price_15min(self) -> float | None:
        return round(sum(self.prices_15min) / len(self.prices_15min), 6) if self.prices_15min else None

    @property
    def all_prices_15min_today(self) -> list[dict]:
        return [{"slot": _quarter_label(i), "price_eur_kwh": p} for i, p in enumerate(self.prices_15min)]

    @property
    def negative_price_slots_today(self) -> list[str]:
        return [_quarter_label(i) for i, p in enumerate(self.prices_15min) if p < 0]

    # ------------------------------------------------------- today -- 15-min buy/sell
    @property
    def current_contract_price_15min(self) -> float | None:
        p = self.current_price_15min
        return self._buy(p) if p is not None else None

    @property
    def current_sell_price_15min(self) -> float | None:
        p = self.current_price_15min
        return self._sell(p) if p is not None else None

    # ------------------------------------------------------- tomorrow -- spot hourly
    @property
    def tomorrow_average_price(self) -> float | None:
        return round(sum(self.prices_tomorrow) / len(self.prices_tomorrow), 6) if self.prices_tomorrow else None

    @property
    def tomorrow_min_price(self) -> float | None:
        src = self.prices_15min_tomorrow or self.prices_tomorrow
        return min(src) if src else None

    @property
    def tomorrow_max_price(self) -> float | None:
        src = self.prices_15min_tomorrow or self.prices_tomorrow
        return max(src) if src else None

    @property
    def tomorrow_cheapest_hours(self) -> list[dict]:
        return [
            {"hour": _hour_label(i), "price_eur_kwh": p}
            for i, p in sorted(enumerate(self.prices_tomorrow), key=lambda x: x[1])
        ]

    @property
    def all_prices_tomorrow(self) -> list[dict]:
        return [{"hour": _hour_label(i), "price_eur_kwh": p} for i, p in enumerate(self.prices_tomorrow)]

    @property
    def negative_price_hours_tomorrow(self) -> list[str]:
        return [_hour_label(i) for i, p in enumerate(self.prices_tomorrow) if p < 0]

    # ------------------------------------------------------- tomorrow -- buy/sell
    @property
    def tomorrow_average_contract_price(self) -> float | None:
        if not self.prices_tomorrow:
            return None
        return round(sum(self._buy(p) for p in self.prices_tomorrow) / len(self.prices_tomorrow), 6)

    @property
    def tomorrow_min_contract_price(self) -> float | None:
        src = self.prices_15min_tomorrow or self.prices_tomorrow
        return self._buy(min(src)) if src else None

    @property
    def tomorrow_max_contract_price(self) -> float | None:
        src = self.prices_15min_tomorrow or self.prices_tomorrow
        return self._buy(max(src)) if src else None

    @property
    def tomorrow_average_sell_price(self) -> float | None:
        if not self.prices_tomorrow:
            return None
        return round(sum(self._sell(p) for p in self.prices_tomorrow) / len(self.prices_tomorrow), 6)

    @property
    def tomorrow_min_sell_price(self) -> float | None:
        src = self.prices_15min_tomorrow or self.prices_tomorrow
        return self._sell(min(src)) if src else None

    @property
    def tomorrow_max_sell_price(self) -> float | None:
        src = self.prices_15min_tomorrow or self.prices_tomorrow
        return self._sell(max(src)) if src else None

    # ------------------------------------------------------- tomorrow -- 15-min
    @property
    def tomorrow_average_price_15min(self) -> float | None:
        return round(sum(self.prices_15min_tomorrow) / len(self.prices_15min_tomorrow), 6) if self.prices_15min_tomorrow else None

    @property
    def tomorrow_min_price_15min(self) -> float | None:
        return min(self.prices_15min_tomorrow) if self.prices_15min_tomorrow else None

    @property
    def tomorrow_max_price_15min(self) -> float | None:
        return max(self.prices_15min_tomorrow) if self.prices_15min_tomorrow else None

    @property
    def all_prices_15min_tomorrow(self) -> list[dict]:
        return [
            {"slot": _quarter_label(i), "price_eur_kwh": p}
            for i, p in enumerate(self.prices_15min_tomorrow)
        ]

    @property
    def negative_price_slots_tomorrow(self) -> list[str]:
        return [_quarter_label(i) for i, p in enumerate(self.prices_15min_tomorrow) if p < 0]

    # ------------------------------------------------------- binary sensor helpers
    @property
    def is_current_price_negative(self) -> bool:
        p = self.current_price
        return p is not None and p < 0

    @property
    def is_current_sell_price_negative(self) -> bool:
        """True when sell price is below SELL_ALERT_THRESHOLD (-0.02 EUR/kWh)."""
        p = self.current_sell_price
        return p is not None and p < SELL_ALERT_THRESHOLD

    @property
    def is_current_price_15min_negative(self) -> bool:
        p = self.current_price_15min
        return p is not None and p < 0

    @property
    def tomorrow_has_negative_prices(self) -> bool:
        return any(p < 0 for p in self.prices_tomorrow)

    @property
    def tomorrow_has_negative_sell_prices(self) -> bool:
        return any(self._sell(p) < SELL_ALERT_THRESHOLD for p in self.prices_tomorrow)


class EngieCoordinator(DataUpdateCoordinator[EngieData]):
    """Coordinator that fetches and parses Engie dynamic energy prices."""

    def __init__(
        self,
        hass: HomeAssistant,
        session,
        buy_multiplier: float = 1.0,
        buy_energy_fee: float = 0.0,
        buy_distribution_fee: float = 0.0,
        sell_fee: float = 0.0,
    ) -> None:
        super().__init__(hass, _LOGGER, name=DOMAIN, update_interval=SCAN_INTERVAL)
        self._session = session
        self._buy_multiplier = buy_multiplier
        self._buy_energy_fee = buy_energy_fee
        self._buy_distribution_fee = buy_distribution_fee
        self._sell_fee = sell_fee
        # Persistent price storage: the XLSX only ever contains ONE day's data.
        # Before 14:00 it has today; after 14:00 it switches to tomorrow.
        # We keep both sets so today's prices remain available after the switch.
        self._today_15min: list[float] = []
        self._tomorrow_15min: list[float] = []
        self._data_date: Optional[date] = None  # date for which _today_15min is valid

    async def _async_update_data(self) -> EngieData:
        today = date.today()
        now = datetime.now()

        # Roll over to a new day: clear both caches.
        if self._data_date != today:
            _LOGGER.debug("New day %s -- clearing cached prices", today)
            self._today_15min = []
            self._tomorrow_15min = []
            self._data_date = today

        # Only fetch when data we actually need is missing.
        need_today = not self._today_15min
        need_tomorrow = now.hour >= AFTERNOON_HOUR and not self._tomorrow_15min

        if need_today or need_tomorrow:
            _LOGGER.debug(
                "Fetching XLSX (need_today=%s, need_tomorrow=%s, hour=%02d)",
                need_today, need_tomorrow, now.hour,
            )
            try:
                async with self._session.get(XLSX_URL) as resp:
                    resp.raise_for_status()
                    content = await resp.read()
            except Exception as err:
                if self._today_15min:
                    _LOGGER.warning("Fetch failed, using cached prices: %s", err)
                else:
                    raise UpdateFailed(f"Error fetching Engie prices: {err}") from err
            else:
                try:
                    file_date, prices = await self.hass.async_add_executor_job(
                        EngieCoordinator._parse_xlsx, content
                    )
                except Exception as err:
                    raise UpdateFailed(f"Error parsing Engie XLSX: {err}") from err

                tomorrow = today + timedelta(days=1)
                if file_date == today:
                    _LOGGER.debug(
                        "XLSX date = today (%s) -- storing %d slots as today's prices",
                        today, len(prices),
                    )
                    self._today_15min = prices
                    if now.hour >= AFTERNOON_HOUR:
                        _LOGGER.warning(
                            "After %02dh00 but XLSX still has today's date (%s) -- "
                            "tomorrow's day-ahead prices not yet published.",
                            AFTERNOON_HOUR, file_date,
                        )
                elif file_date == tomorrow:
                    _LOGGER.debug(
                        "XLSX date = tomorrow (%s) -- storing %d slots as tomorrow's prices",
                        tomorrow, len(prices),
                    )
                    self._tomorrow_15min = prices
                else:
                    _LOGGER.warning(
                        "Unexpected date in XLSX: %s (today=%s, tomorrow=%s) -- ignoring.",
                        file_date, today, tomorrow,
                    )
        else:
            _LOGGER.debug(
                "Using cached prices: today=%d slots, tomorrow=%d slots",
                len(self._today_15min), len(self._tomorrow_15min),
            )

        return EngieData(
            price_date=today,
            prices=_to_hourly(self._today_15min),
            prices_tomorrow=_to_hourly(self._tomorrow_15min),
            prices_15min=list(self._today_15min),
            prices_15min_tomorrow=list(self._tomorrow_15min),
            buy_multiplier=self._buy_multiplier,
            buy_energy_fee=self._buy_energy_fee,
            buy_distribution_fee=self._buy_distribution_fee,
            sell_fee=self._sell_fee,
        )

    @staticmethod
    def _parse_xlsx(content: bytes) -> tuple[date, list[float]]:
        """Parse the Kwartieren/Quartiers sheet.

        Returns (price_date, prices) where prices is a list of up to 96 EUR/kWh values.

        Sheet layout:
          - B1/B2: merged cell containing the date these prices are valid for ('geldig voor').
          - Col A: quarter-hour time labels like '00:00 - 00:15'.
          - Col B: spot price in EUR/MWh (e.g. 140.61 = 0.14061 EUR/kWh).
        """
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        _LOGGER.debug("XLSX sheets: %s", [ws.title for ws in wb.worksheets])

        # Find the Kwartieren / Quartiers sheet.
        sheet = next(
            (
                ws for ws in wb.worksheets
                if any(
                    kw in ws.title
                    for kw in ("Kwartier", "kwartier", "Quartier", "quartier", "Quart", "quart")
                )
            ),
            wb.worksheets[0],
        )
        _LOGGER.debug("Using sheet: '%s'", sheet.title)

        # Date is in merged cell B1/B2 (rows 1-2, column 2).
        # openpyxl stores merged-cell values only in the top-left cell of the merge.
        date_val = sheet.cell(row=1, column=2).value
        if date_val is None:
            date_val = sheet.cell(row=2, column=2).value

        if isinstance(date_val, datetime):
            price_date = date_val.date()
        elif isinstance(date_val, date):
            price_date = date_val
        else:
            _LOGGER.warning(
                "Could not read date from merged cell B1/B2 (got %r) -- defaulting to today.",
                date_val,
            )
            price_date = date.today()

        _LOGGER.debug("Sheet 'geldig voor': %s", price_date)

        # Read column B (index 1): EUR/MWh values, convert to EUR/kWh.
        # Skip any non-numeric rows (headers, labels, empty rows).
        # Stop once we have 96 slots.
        prices: list[float] = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if len(prices) >= SLOTS_15MIN:
                break
            if len(row) < 2:
                continue
            val = row[1]  # column B
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                prices.append(_to_kwh(float(val)))

        _LOGGER.debug(
            "Read %d quarter-hour prices from sheet '%s'",
            len(prices), sheet.title,
        )
        return price_date, prices
